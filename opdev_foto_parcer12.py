import asyncio   
import time
from playwright.async_api import async_playwright
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook  # Используем load_workbook для загрузки существующего файла

# Твои данные
LOGIN_URL = 'https://opdev.ru/'
TARGET_URL = 'https://opdev.ru/catalog/opticheskie_nablyudatelnye_pribory/pritsely_dnevnye_kollimatornye/pritsely_dnevnye_brite/pritsel_opticheskiy_brite_wa8x_2-16x50_sf_ir/'

USER_LOGIN = '9896600@mail.ru'
USER_PASSWORD = '89111664848'
TARGET_ARTICLE = '31509'  # Заданный артикул для поиска

# Функция для отсчета времени с обновлением в реальном времени
async def countdown_timer(step_name, duration, stop_event):
    start_time = time.time()
    print(step_name)  # Печатаем описание этапа один раз
    while not stop_event.is_set():  # Проверяем, если остановка не активирована
        elapsed_time = time.time() - start_time
        print(f"{int(elapsed_time)} секунд прошло", end="\r", flush=True)  # Обновляем только цифры секунд
        await asyncio.sleep(1)
    print(f"{int(elapsed_time)} секунд прошло")  # Завершаем вывод с временем выполнения этапа

async def run():
    start_time = time.time()  # Засекаем время начала выполнения скрипта

    print("Привет! Запускаем парсер через Playwright...")

    # Проверяем, существует ли файл Excel. Если да, открываем, если нет — создаем новый
    excel_file = os.path.join(os.path.dirname(__file__), "results.xlsx")
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)  # Загружаем существующий файл
        ws = wb.active
    else:
        wb = Workbook()  # Создаем новый файл
        ws = wb.active
        ws.title = "Товары"
        # Записываем заголовки в Excel
        ws.append(["Наименование", "Значение"])

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)  # headless=True если нужно без окна
        page = await browser.new_page()

        # Этап 1: Ожидаем открытия сайта и ждем авторизацию
        stop_event = asyncio.Event()  # Ожидаем события остановки
        timer_task = asyncio.create_task(countdown_timer("Открыли сайт\nЖдем авторизацию...", 30, stop_event))  # Запускаем таймер в фоновом потоке
        await page.goto(LOGIN_URL)  # Основной шаг (переход на сайт)
        stop_event.set()  # Завершаем таймер сразу, когда страница открыта
        await timer_task  # Ждем завершения таймера

        # Этап 2: Вводим логин и пароль
        await page.wait_for_selector('input[name="USER_LOGIN"]')
        await page.wait_for_selector('input[name="USER_PASSWORD"]')
        await page.fill('input[name="USER_LOGIN"]', USER_LOGIN)
        await page.fill('input[name="USER_PASSWORD"]', USER_PASSWORD)
        await page.get_by_role("button", name="Войти").click()

        # Этап 3: Ожидаем успешную авторизацию
        stop_event.clear()  # Сбросим стоп-сигнал для следующего этапа
        timer_task = asyncio.create_task(countdown_timer("Успешно авторизовались!\nОткрываем страницу товара...", 30, stop_event))
        await page.wait_for_load_state('networkidle')
        stop_event.set()  # Завершаем таймер сразу, когда страница загрузилась
        await timer_task

        # Этап 4: Используем поиск по артикулу
        search_input = await page.query_selector('#inheader-title-search-input')  # Поиск по артикулу
        await search_input.fill(TARGET_ARTICLE)  # Вводим артикул
        await page.wait_for_timeout(1000)  # Подождем, пока начнется поиск (крутящийся индикатор)

        # Ждем появления подсказки после того, как она загрузится
        await page.wait_for_selector('.kt-quick-search__item-wrapper', timeout=5000)  # Ожидаем, что элементы подсказки появятся

        # Ищем предложенный товар с артикулом 31509
        suggestions = await page.query_selector_all('.kt-quick-search__item-wrapper')  # Все элементы подсказки
        found = False
        
        for suggestion in suggestions:
            article_text = await suggestion.query_selector('.kt-quick-search__item-desc')
            if article_text:
                article_text_value = await article_text.inner_text()
                if TARGET_ARTICLE in article_text_value:  # Если артикул найден в подсказке
                    link = await suggestion.query_selector('a')  # Получаем ссылку
                    if link:
                        print(f"Найден товар с артикулом {TARGET_ARTICLE}. Кликаем на него...")
                        await link.click()  # Кликаем по предложенному товару
                        found = True
                        break

        if not found:
            print(f"Товар с артикулом {TARGET_ARTICLE} не найден в подсказке.")

        # Ждем полной загрузки страницы товара
        await page.wait_for_load_state('networkidle')  # Ждем, пока все элементы страницы загрузятся

        # Этап 5: Парсим название товара и артикул
        product_name_element = await page.query_selector('div.mb-2 span')  # Первый span, который содержит название товара
        article_element = await page.query_selector('div span:has-text("Арт:")')  # Второй span, который содержит "Арт:"

        if product_name_element and article_element:
            product_name = await product_name_element.inner_text()  # Название товара
            article_text = await article_element.inner_text()  # Получаем текст из элемента
            article = article_text.replace('Арт:', '').strip()  # Артикул без "Арт:" и лишних пробелов

            print(f"Название товара: {product_name}")
            print(f"Артикул: {article}")
        else:
            product_name, article = "Не указано", "Не указано"

        # Этап 6: Парсим таблицу с характеристиками
        characteristics = {}
        rows = await page.query_selector_all('table tbody tr')
        for row in rows:
            cols = await row.query_selector_all('td')
            if len(cols) == 2:
                label = await cols[0].inner_text()  # Получаем текст из первой ячейки
                value = await cols[1].inner_text()  # Получаем текст из второй ячейки
                characteristics[label.strip()] = value.strip()  # Добавляем пару в словарь

        # Вставляем две пустые строки перед новой записью
        ws.append(["", ""])  # Вставляем первую пустую строку
        ws.append(["", ""])  # Вставляем вторую пустую строку

        # Записываем данные в Excel
        ws.append(["Наименование товара", product_name])
        ws.append(["Артикул", article])

        for label, value in characteristics.items():
            print(f"{label}: {value}")  # Выводим в консоль
            ws.append([label, value])

        # Ищем все кнопки с картинками
        buttons = await page.query_selector_all('.product-images-thumbs button')

        image_urls = []
        if not buttons:
            print("Не нашли кнопки с картинками 😢")
        else:
            print(f"Нашли {len(buttons)} изображений:")

            # Сохраняем результаты в список
            for button in buttons:
                style = await button.get_attribute('style')
                if style:
                    # Вытащим URL из стиля background-image
                    import re
                    match = re.search(r'url\(["\']?(.*?)["\']?\)', style)
                    if match:
                        relative_url = match.group(1)
                        full_url = 'https://opdev.ru' + relative_url
                        image_urls.append(full_url)

        # Записываем ссылки на изображения в Excel
        ws.append(["Ссылки на фото", ", ".join(image_urls)])

        # Выводим список ссылок на фото в консоль
        print("\nСсылки на фото:")
        for url in image_urls:
            print(url)

        # Замеряем общее время выполнения
        total_time = time.time() - start_time
        print(f"\nЗатрачено секунд на выполнение всех этапов: {total_time:.0f}  🙂")

        # Сохраняем Excel файл
        wb.save(os.path.join(os.path.dirname(__file__), "results.xlsx"))
        print("Результаты записаны в Excel файл: results.xlsx")

        # Закрываем браузер
        await browser.close()

# Начинаем выполнение
asyncio.run(run())
